# -*- coding: utf-8 -*-
"""
Created on Sun Nov 22 17:23:00 2020

@author: Victor
"""

import pandas as pd
from datetime import timedelta, datetime
from openpyxl import load_workbook

def get_date_ref(filename,sheet_name):
    date_ref = datetime.now() - timedelta(days=11);
    date_ref = date_ref.date();
    try:
        writer = pd.ExcelWriter(filename,engine="openpyxl");
        writer.book = load_workbook(filename);
        maxrow = writer.book[sheet_name].max_row;
        date_ref = writer.book[sheet_name]["A"+str(maxrow)].value.date();
    except FileNotFoundError:
        print('File not found : reference date set to 7 days ago -->'+ str(date_ref));
    except KeyError:
        print('File not found : reference date set to 10 days ago -->'+ str(date_ref));
    return date_ref;

def get_titre_ref(filename,sheet_name):
    
    titre_ref = "";
    
    try:
        writer =  pd.ExcelWriter(filename,engine="openpyxl");
        writer.book = load_workbook(filename);
        maxrow = writer.book[sheet_name].max_row;
        titre_ref = writer.book[sheet_name]["B"+str(maxrow)].value;
        # print('titre_ref : {}'.format(titre_ref))
    except FileNotFoundError:
        print('File not found : getting all ads');
    except KeyError:
        print('Sheet not found : getting all ads');
        
    return titre_ref;

def parse_and_print(filename,company_class,ws):
    
    data = pd.DataFrame();
    
    parsing = company_class.parsing_mode;
    
    if parsing == "date" or parsing == "title":
    
        if parsing == 'date':
            
            print("{} parsing from date ".format(company_class.name));
            
            date_ref = get_date_ref(filename, company_class.name);
            data = ws.parse_from_date(date_ref)
            
        elif parsing == 'title':
            
            print("{} parsing from title ".format(company_class.name));
            titre_ref = get_titre_ref(filename, company_class.name);
            data = ws.parse_from_title(titre_ref);

        ws.print_to_excel(filename, data, sheet_name = company_class.name, header = False);
    
    else:
        print("{} parsing mode not found".format(parsing));
    
    return data;