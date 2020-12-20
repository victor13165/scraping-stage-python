# -*- coding: utf-8 -*-
"""
Created on Sun Nov 22 12:01:31 2020

@author: Victor Baconnet
"""

import pandas as pd
from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium import webdriver


class WSInternship():
    
    #Constructeur
    def __init__(self, company_class):
        
        self.log_file_name = "new_internships_log.txt";
        self.company_class = company_class; #Objet Safran, Volvo... 
        self.name = company_class.name; #Nom entreprise
        
        self.data = {"Date":[],"Titre":[],"Url":[]}; #Va stocker date, titre et url de l'annonce
        self.url = company_class.url; #url de départ
        
        self.annonces =  [];
        
        mode = self.company_class.scrapping_mode;
        print(self.name + ": requesting "+self.url);
        
        if mode == 'bs': #Requesting with beautifulsoup
            self.annonces = self.company_class.bs_procedure()
        
        elif mode == 'selenium': #Requesting with selenium
            options = FirefoxOptions()
            options.add_argument("--headless")
            print("Opening webdriver")
            driver = webdriver.Firefox(options=options)
            try:
                self.annonces = self.company_class.selenium_procedure(driver);
            except:
                print("{} : Unable to retrieve annonces".format(self.company_class.name));
                
    
    #Pour obtenir la liste du paramètre demandé (date, titre ou url)
    def get_param(self,param):
        return self.data[param];
    
    #Retourne un DataFrame avec les annonces triées du moins récent au plus récent
    def parse_from_date(self,date_ref):
        
        for annonce in self.annonces: #pour chaque annonce
            
            date = self.company_class.get_date(annonce); #Recup la date de post de l'annonce
            if date > date_ref: #Si la date de publication est plus récente 
                
                self.data["Date"].append(date); #Ajouter la date
                self.data["Url"].append(self.company_class.get_url(annonce)); #Ajouter l'url
                self.data["Titre"].append(self.company_class.get_titre(annonce)); #Et le titre
            
            else: #Dans ce cas là, il n'y a plus de nouvelles annonces
                print("Done Scraping "+self.name);
                break; 
            
        try:
            self.company_class.close_driver();
        except:
            pass;
        data_pd = pd.DataFrame(self.data).set_index("Date");
        data_pd = data_pd.sort_index(ascending=True);
        return data_pd;
    
    #Retourne un dataframe pour les sites où la date de publication n'apparaît pas
    def parse_from_title(self,titre_ref):
        # index = 0;
        for annonce in self.annonces: #Pour chaque annonce
            annonce.location_once_scrolled_into_view;
            titre = self.company_class.get_titre(annonce); #Titre de l'annonce
            # print(index)
            # index += 1
            if (titre != titre_ref): #On compare avec un titre de référence
                
                self.data["Date"].append(self.company_class.get_date(annonce))
                self.data["Url"].append(self.company_class.get_url(annonce));
                self.data["Titre"].append(self.company_class.get_titre(annonce));
                
            else:
                break;
        
        print("Done scraping "+self.name);
        try:
            self.company_class.close_driver();
        except:
            pass;
        data_pd = pd.DataFrame(self.data);
        data_pd = data_pd.sort_index(ascending=False);
        data_pd = data_pd.set_index("Date");
        return data_pd;
                
    
    def print_to_excel(self, filename, df, sheet_name, startrow=None, **to_excel_kwargs):
        
        # print('okok');
        # log_file = open("new_internships_log.txt","w");
        # log_file.write(self.name + ": " + str(len(df)) + " nouvelles offres")
        
        if len(df) != 0: #Si il y a des nouvelles offres
        
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            
            try:
                # try to open an existing workbook
                writer.book = load_workbook(filename)
                
                # get the last row in the existing Excel sheet
                # if it was not specified explicitly
                if startrow is None and sheet_name in writer.book.sheetnames:
                    startrow = writer.book[sheet_name].max_row
                 
                # copy existing sheets
                writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
            
            except FileNotFoundError:
                # file does not exist yet, we will create it
                pass
            
            if startrow is None:
                startrow = 0
                if 'header' in to_excel_kwargs:
                    to_excel_kwargs.pop('header'); #Dans ce cas là on veut le header puisque pas de start_row
            
            
            # write out the new sheet
            df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
            
            # save the workbook
            writer.save()
            
        
            
        
        